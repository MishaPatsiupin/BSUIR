"""
Patched PySide6 client that verifies admin password before allowing add/update/delete
on lookup tables. Uses /api/check_admin endpoint.
Requirements: PySide6, requests
Run: python qt_client_patch.py
"""

import sys
import json
import csv
import traceback
from urllib.parse import urlencode
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import requests
from PySide6.QtCore import (Qt, QAbstractTableModel, QModelIndex, QObject,
                            Signal, Slot, QRunnable, QThreadPool)
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                               QHBoxLayout, QTableView, QPushButton, QToolBar,
                               QComboBox, QFileDialog, QMessageBox, QDialog,
                               QLabel, QLineEdit, QFormLayout, QStatusBar,
                               QInputDialog)
from PySide6.QtGui import QKeySequence, QShortcut, QAction

# --------------------- Worker utilities ---------------------
class WorkerSignals(QObject):
    finished = Signal(object, object)  # (result, error)

class Worker(QRunnable):
    def __init__(self, fn, *args, **kwargs):
        super().__init__()
        self.fn = fn
        self.args = args
        self.kwargs = kwargs
        self.signals = WorkerSignals()

    def run(self):
        try:
            res = self.fn(*self.args, **self.kwargs)
            self.signals.finished.emit(res, None)
        except Exception as e:
            # Передаём только сообщение ошибки, без стека
            self.signals.finished.emit(None, str(e))

# --------------------- ApiClient (blocking requests in thread) ---------------------
class ApiClient(QObject):
    request_finished = Signal(object, object)  # (result, error)

    def __init__(self, base_url="http://127.0.0.1:8000"):
        super().__init__()
        self.base_url = base_url.rstrip('/')
        self.pool = QThreadPool.globalInstance()
        self.superpass = None  # admin password for backup/lookup ops

    def set_superpass(self, pw: str):
        self.superpass = pw

    def clear_superpass(self):
        self.superpass = None

    def _url(self, path):
        return f"{self.base_url}/api/{path}"

    def _do_request(self, method, path, json_data=None, params=None):
        url = self._url(path)
        if params:
            url = url + "?" + urlencode(params)
        headers = {"Content-Type": "application/json; charset=utf-8"}
        if self.superpass:
            headers["X-Superpass"] = self.superpass

        try:
            r = requests.request(method, url, json=json_data, headers=headers, timeout=10)
            r.raise_for_status()
        except requests.exceptions.HTTPError as e:
            # человекочитаемое сообщение с кодом и телом ответа
            body = ""
            try:
                body = r.json()
            except Exception:
                body = r.text
            raise Exception(f"HTTP {r.status_code}: {body}")
        except requests.exceptions.RequestException as e:
            raise Exception(f"Network error: {e}")

        text = r.text.strip()
        if not text:
            return None
        try:
            return json.loads(text)
        except Exception:
            try:
                return json.loads(text.encode('utf-8', 'surrogatepass').decode('utf-8-sig'))
            except Exception:
                return text


    def _on_worker_finished(self, result, error):
        print(f"[ApiClient] worker finished: result_type={type(result).__name__} error={bool(error)}")
        self.request_finished.emit(result, error)

    def request(self, method, path, json_data=None, params=None):
        worker = Worker(self._do_request, method, path, json_data, params)
        worker.signals.finished.connect(self._on_worker_finished)
        self.pool.start(worker)

    # convenience methods
    def get_table(self, table_key, params=None):
        self.request('GET', table_key, None, params)

    def post(self, table_key, data):
        self.request('POST', table_key, data)

    def put(self, table_key, data):
        self.request('PUT', table_key, data)

    def delete(self, table_key, data):
        self.request('DELETE', table_key, data)

    def backup(self):
        self.request('GET', 'backup')

    def restore(self):
        self.request('POST', 'restore')

    # verify admin by posting to /api/check_admin
    def verify_superpass(self, password, callback):
        """
        callback(result, error) will be called in main thread.
        """
        worker = Worker(self._do_request, 'POST', 'check_admin', {'password': password}, None)
        worker.signals.finished.connect(callback)
        self.pool.start(worker)

# --------------------- Table model ---------------------
class TableModel(QAbstractTableModel):
    def __init__(self, columns=None, rows=None):
        super().__init__()
        self.columns = columns or []
        self.rows = rows or []

    def rowCount(self, parent=QModelIndex()):
        return len(self.rows)

    def columnCount(self, parent=QModelIndex()):
        return len(self.columns)

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None
        if role in (Qt.ItemDataRole.DisplayRole, Qt.ItemDataRole.EditRole):
            row = self.rows[index.row()]
            col = self.columns[index.column()]
            val = row.get(col, "")
            return str(val)
        return None

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role != Qt.ItemDataRole.DisplayRole:
            return None
        if orientation == Qt.Orientation.Horizontal:
            return self.columns[section]
        return str(section + 1)

    def update(self, columns, rows):
        self.beginResetModel()
        self.columns = columns
        self.rows = rows
        self.endResetModel()

# --------------------- Dynamic dialog for Add/Edit ---------------------
class RecordDialog(QDialog):
    def __init__(self, columns, data=None, parent=None, readonly=None, exclude=None):
        super().__init__(parent)
        self.setWindowTitle('Edit record' if data else 'Add record')
        self.columns = [c for c in columns if not (exclude and c in exclude)]
        self.data = data or {}
        self.fields = {}
        layout = QFormLayout()
        for col in self.columns:
            le = QLineEdit()
            le.setText(str(self.data.get(col, '')))
            if readonly and col in readonly:
                le.setReadOnly(True)
            layout.addRow(QLabel(col), le)
            self.fields[col] = le
        btn_ok = QPushButton('OK')
        btn_cancel = QPushButton('Cancel')
        btn_ok.clicked.connect(self.accept)
        btn_cancel.clicked.connect(self.reject)
        hl = QHBoxLayout()
        hl.addWidget(btn_ok)
        hl.addWidget(btn_cancel)
        layout.addRow(hl)
        self.setLayout(layout)
        self.setModal(True)
        self.setFixedSize(self.sizeHint())

    def get_data(self):
        return {col: self.fields[col].text() for col in self.columns}

# --------------------- Main Window ---------------------
class MainWindow(QMainWindow):
    TABLE_KEYS = [
        'patients', 'appointments', 'billing', 'departments', 'doctors', 'rooms',
        'doctor_department', 'doctor_patient', 'patient_room', 'medical_records'
    ]
    LOOKUP_KEYS = {'departments', 'rooms'}

    def __init__(self):
        super().__init__()
        self.setWindowTitle('Hospital Admin Client (patched)')
        self.resize(1000, 600)
        self.api = ApiClient('http://127.0.0.1:8000')
        self.api.request_finished.connect(self.on_api_result)
        self.last_result = None
        self.current_table = None

        # ==== меню ====
        menubar = self.menuBar()
        file_menu = menubar.addMenu("&File")   # Alt+F откроет меню

        act_export = QAction("&Export CSV...", self)
        act_export.setShortcut("Ctrl+Shift+E")
        act_export.triggered.connect(self.on_export)
        file_menu.addAction(act_export)

        file_menu.addSeparator()

        act_exit = QAction("E&xit", self)
        act_exit.setShortcut("Ctrl+E")
        act_exit.triggered.connect(self.close)
        file_menu.addAction(act_exit)

        # ==== тулбар ====
        toolbar = QToolBar()
        self.addToolBar(toolbar)

        self.cmb_tables = QComboBox()
        self.cmb_tables.addItems(self.TABLE_KEYS)
        self.cmb_tables.currentTextChanged.connect(self.on_table_change)
        toolbar.addWidget(self.cmb_tables)

        btn_reload = QPushButton('Reload')
        btn_reload.clicked.connect(lambda: self.load_table(self.current_table))
        toolbar.addWidget(btn_reload)

        btn_add = QPushButton('Add')
        btn_add.clicked.connect(self.on_add)
        toolbar.addWidget(btn_add)

        btn_edit = QPushButton('Edit')
        btn_edit.clicked.connect(self.on_update)
        toolbar.addWidget(btn_edit)

        btn_delete = QPushButton('Delete')
        btn_delete.clicked.connect(self.on_delete)
        toolbar.addWidget(btn_delete)

        btn_backup = QPushButton('Backup')
        btn_backup.clicked.connect(self.on_backup)
        toolbar.addWidget(btn_backup)

        btn_restore = QPushButton('Restore')
        btn_restore.clicked.connect(self.on_restore)
        toolbar.addWidget(btn_restore)

        btn_export = QPushButton('Export')
        btn_export.clicked.connect(self.on_export)
        toolbar.addWidget(btn_export)

        btn_query = QPushButton('Special Query')
        btn_query.clicked.connect(self.on_query)
        toolbar.addWidget(btn_query)

        btn_query = QPushButton('Save table')
        btn_query.clicked.connect(self.on_save_current_table)
        toolbar.addWidget(btn_query)

        # ==== центральный виджет ====
        central = QWidget()
        self.setCentralWidget(central)
        v = QVBoxLayout(central)

        self.table_view = QTableView()
        self.model = TableModel()
        self.table_view.setModel(self.model)
        self.table_view.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
        v.addWidget(self.table_view)

        self.status = QStatusBar()
        self.setStatusBar(self.status)

        self.setup_shortcuts()

        # default table
        self.cmb_tables.setCurrentText('appointments')
        self.current_table = 'appointments'
        self.load_table(self.current_table)

    def setup_shortcuts(self):
        QShortcut(QKeySequence('Ctrl+A'), self, activated=self.on_add)
        QShortcut(QKeySequence('Ctrl+V'), self, activated=lambda: self.load_table(self.current_table))
        QShortcut(QKeySequence('Ctrl+D'), self, activated=self.on_delete)
        QShortcut(QKeySequence('Ctrl+U'), self, activated=self.on_update)
        QShortcut(QKeySequence('Ctrl+Q'), self, activated=self.on_query)
        QShortcut(QKeySequence('Ctrl+S'), self, activated=self.on_save_current_table)
        QShortcut(QKeySequence('Ctrl+B'), self, activated=self.on_backup)
        QShortcut(QKeySequence('Ctrl+E'), self, activated=self.close)

    def on_save_current_table(self):
        if not self.current_table or not self.last_result:
            QMessageBox.warning(self, "Экспорт", "Нет данных для сохранения")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = self.current_table

        headers = list(self.last_result[0].keys())
        ws.append(headers)

        for row in self.last_result:
            ws.append([row.get(h) for h in headers])

        for col_idx, col in enumerate(headers, 1):
            max_len = max((len(str(r.get(col) or "")) for r in self.last_result), default=len(col))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

        path = f"{self.current_table}.xlsx"
        wb.save(path)

        QMessageBox.information(self, "Экспорт", f"Таблица сохранена в {path}")

    def on_table_change(self, key):
        self.current_table = key
        self.load_table(key)

    def load_table(self, key, params=None):
        if not key:
            return
        self.status.showMessage(f'Loading {key}...')
        print(f"[MainWindow] load_table: {key} params={params}")
        self.api.get_table(key, params=params)

    @Slot(object, object)
    def on_api_result(self, result, error):
        if error:
            msg = str(error).strip()
            print(f"[MainWindow] API error: {msg}")
            QMessageBox.critical(self, 'Ошибка API', msg)
            self.status.showMessage('Ошибка')
            return

        # дальше твой код как раньше
        if isinstance(result, list):
            self.last_result = result
            columns = list(result[0].keys()) if result else []
            self.model.update(columns, result)
            self.status.showMessage(f'Загружено {len(result)} строк')
        elif isinstance(result, dict):
            self.last_result = result
            if result.get('status') == 'ok':
                self.status.showMessage('Операция успешна')
                if self.current_table:
                    self.load_table(self.current_table)
            else:
                msg = json.dumps(result, ensure_ascii=False)
                self.status.showMessage(msg)
                QMessageBox.warning(self, 'Ответ сервера', msg)
        elif isinstance(result, str):
            self.status.showMessage(result)
            QMessageBox.information(self, 'Ответ сервера', result)
        elif result is None:
            self.status.showMessage('Пустой ответ')
        else:
            self.status.showMessage('Неизвестный ответ')


    def selected_row(self):
        sel = self.table_view.selectionModel().selectedRows()
        if not sel:
            return None, None
        row_idx = sel[0].row()
        row = self.model.rows[row_idx]
        return row_idx, row

    def _find_pk(self, row):
        for k in row.keys():
            if k.endswith('_id'):
                return k
        return None

    # asynchronous flow: prompt -> verify via /api/check_admin -> on_success()
    def ask_and_verify_admin(self, on_success):
        pw, ok = QInputDialog.getText(self, 'Admin required', 'Enter admin password:', QLineEdit.Password)
        if not ok or not pw:
            return  # cancelled

        self.status.showMessage('Verifying admin password...')

        def verify_callback(result, error):
            # runs in main thread
            if error:
                QMessageBox.critical(self, 'Verification error', str(error))
                self.status.showMessage('Verification failed')
                return
            if isinstance(result, dict) and result.get('ok'):
                self.api.set_superpass(pw)
                self.status.showMessage('Admin verified')
                try:
                    on_success()
                except Exception as e:
                    print('on_success handler error:', e)
            else:
                msg = result.get('error') if isinstance(result, dict) else 'Invalid password'
                QMessageBox.warning(self, 'Invalid password', str(msg))
                self.status.showMessage('Invalid admin password')

        self.api.verify_superpass(pw, verify_callback)

    def on_add(self):
        def do_add():
            pk_candidates = [c for c in (self.model.columns or []) if c.endswith('_id')]
            dlg_columns = [c for c in (self.model.columns or []) if c not in pk_candidates]
            dlg = RecordDialog(dlg_columns, data=None, parent=self, readonly=None, exclude=None)
            if dlg.exec() == QDialog.Accepted:
                data = dlg.get_data()
                self.api.post(self.current_table, data)
                self.status.showMessage('Adding...')

        if self.current_table in self.LOOKUP_KEYS:
            self.ask_and_verify_admin(on_success=do_add)
            return

        do_add()

    def on_update(self):
        idx, row = self.selected_row()
        if row is None:
            QMessageBox.information(self, 'Info', 'Select a row first')
            return

        def do_update():
            pk = self._find_pk(row)
            if pk:
                editable_columns = [c for c in self.model.columns if c != pk]
            else:
                editable_columns = list(self.model.columns)
            dlg = RecordDialog(editable_columns, data=row, parent=self, readonly=None, exclude=None)
            if dlg.exec() == QDialog.Accepted:
                data = dlg.get_data()
                if pk:
                    data[pk] = row[pk]
                self.api.put(self.current_table, data)
                self.status.showMessage('Updating...')

        if self.current_table in self.LOOKUP_KEYS:
            self.ask_and_verify_admin(on_success=do_update)
            return

        do_update()

    def on_delete(self):
        idx, row = self.selected_row()
        if row is None:
            QMessageBox.information(self, 'Info', 'Select a row first')
            return

        def do_delete():
            reply = QMessageBox.question(self, 'Confirm', 'Delete selected record?')
            if reply != QMessageBox.StandardButton.Yes:
                return
            pk = self._find_pk(row)
            if not pk:
                QMessageBox.critical(self, 'Error', 'Primary key not found')
                return
            payload = {pk: row[pk]}
            self.api.delete(self.current_table, payload)
            self.status.showMessage('Deleting...')

        if self.current_table in self.LOOKUP_KEYS:
            self.ask_and_verify_admin(on_success=do_delete)
            return

        do_delete()

    def on_query(self):
        k, ok = QInputDialog.getText(self, 'Query', 'Enter filter as key=value (e.g. doctor_id=5):')
        if not ok or not k:
            return
        try:
            key, val = k.split('=', 1)
        except Exception:
            QMessageBox.warning(self, 'Bad format', 'Use key=value')
            return
        self.load_table(self.current_table, params={key: val})

    def on_export(self):
        if self.last_result is None:
            QMessageBox.information(self, 'Info', 'Nothing to export')
            return
        rows = self.last_result if isinstance(self.last_result, list) else []
        if not rows:
            QMessageBox.information(self, 'Info', 'No rows to save')
            return
        path, _ = QFileDialog.getSaveFileName(self, 'Save CSV', '', 'CSV files (*.csv)')
        if not path:
            return
        cols = list(rows[0].keys())
        try:
            with open(path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=cols)
                writer.writeheader()
                for r in rows:
                    writer.writerow(r)
            self.status.showMessage(f'Exported {len(rows)} rows to {path}')
        except Exception as e:
            QMessageBox.critical(self, 'Error', str(e))

    def on_backup(self):
        reply = QMessageBox.question(self, 'Confirm', 'Create backup now?')
        if reply != QMessageBox.StandardButton.Yes:
            return
        if not self._prompt_superpass_and_set():
            QMessageBox.warning(self, 'Permission', 'Admin password required for backup')
            return
        self.api.backup()
        self.status.showMessage('Backup started...')

    def _prompt_superpass_and_set(self):
        pw, ok = QInputDialog.getText(self, 'Admin required', 'Enter admin password:', QLineEdit.Password)
        if not ok or not pw:
            return False
        self.api.set_superpass(pw)
        return True

    def on_restore(self):
        reply = QMessageBox.question(self, 'Confirm', 'Restore database from backup? This is destructive.')
        if reply != QMessageBox.StandardButton.Yes:
            return
        if not self._prompt_superpass_and_set():
            QMessageBox.warning(self, 'Permission', 'Admin password required for restore')
            return
        self.api.restore()
        self.status.showMessage('Restore started...')

# --------------------- Entry point ---------------------
if __name__ == '__main__':
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec())
