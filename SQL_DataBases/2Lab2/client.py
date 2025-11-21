import sys
import json
import csv
import traceback
from urllib.parse import urlencode
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import requests
from PySide6.QtCore import (Qt, QAbstractTableModel, QModelIndex, QObject,
                            Signal, Slot, QRunnable, QThreadPool, QSize)
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                               QHBoxLayout, QTableView, QPushButton, QToolBar,
                               QComboBox, QFileDialog, QMessageBox, QDialog,
                               QLabel, QLineEdit, QFormLayout, QStatusBar,
                               QInputDialog, QStyle, QHeaderView, QGridLayout,
                               QFrame)
from PySide6.QtGui import QKeySequence, QShortcut, QAction, QPalette, QColor, QFont

# --------------------- Worker utilities ---------------------
class WorkerSignals(QObject):
    finished = Signal(object, object)  # (result, error)

class Worker(QRunnable):
    def __init__(self, fn, *args, **kwargs):
        super().__init__()
        self.fn = fn
        self.args = args
        self.kwargs = kwargs
        self.signals = WorkerSignals()

    def run(self):
        try:
            res = self.fn(*self.args, **self.kwargs)
            self.signals.finished.emit(res, None)
        except Exception as e:
            self.signals.finished.emit(None, str(e))

# --------------------- ApiClient (blocking requests in thread) ---------------------
class ApiClient(QObject):
    request_finished = Signal(object, object)  # (result, error)

    def __init__(self, base_url="http://127.0.0.1:8000"):
        super().__init__()
        self.base_url = base_url.rstrip('/')
        self.pool = QThreadPool.globalInstance()
        self.superpass = None

    def set_superpass(self, pw: str):
        self.superpass = pw

    def clear_superpass(self):
        self.superpass = None

    def _url(self, path):
        return f"{self.base_url}/api/{path}"

    def _do_request(self, method, path, json_data=None, params=None):
        url = self._url(path)
        if params:
            url = url + "?" + urlencode(params)
        headers = {"Content-Type": "application/json; charset=utf-8"}
        if self.superpass:
            headers["X-Superpass"] = self.superpass

        try:
            r = requests.request(method, url, json=json_data, headers=headers, timeout=10)
            r.raise_for_status()
        except requests.exceptions.HTTPError as e:
            body = ""
            try:
                body = r.json()
            except Exception:
                body = r.text
            raise Exception(f"HTTP {r.status_code}: {body}")
        except requests.exceptions.RequestException as e:
            raise Exception(f"Network error: {e}")

        text = r.text.strip()
        if not text:
            return None
        try:
            return json.loads(text)
        except Exception:
            try:
                return json.loads(text.encode('utf-8', 'surrogatepass').decode('utf-8-sig'))
            except Exception:
                return text

    def _on_worker_finished(self, result, error):
        self.request_finished.emit(result, error)

    def request(self, method, path, json_data=None, params=None):
        worker = Worker(self._do_request, method, path, json_data, params)
        worker.signals.finished.connect(self._on_worker_finished)
        self.pool.start(worker)

    def get_table(self, table_key, params=None):
        self.request('GET', table_key, None, params)

    def post(self, table_key, data):
        self.request('POST', table_key, data)

    def put(self, table_key, data):
        self.request('PUT', table_key, data)

    def delete(self, table_key, data):
        self.request('DELETE', table_key, data)

    def backup(self):
        self.request('GET', 'backup')

    def restore(self):
        self.request('POST', 'restore')

    def verify_superpass(self, password, callback):
        worker = Worker(self._do_request, 'POST', 'check_admin', {'password': password}, None)
        worker.signals.finished.connect(callback)
        self.pool.start(worker)

# --------------------- Table model ---------------------
class TableModel(QAbstractTableModel):
    def __init__(self, columns=None, rows=None):
        super().__init__()
        self.columns = columns or []
        self.rows = rows or []

    def rowCount(self, parent=QModelIndex()):
        return len(self.rows)

    def columnCount(self, parent=QModelIndex()):
        return len(self.columns)

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None
        if role in (Qt.ItemDataRole.DisplayRole, Qt.ItemDataRole.EditRole):
            row = self.rows[index.row()]
            col = self.columns[index.column()]
            val = row.get(col, "")
            return str(val)
        return None

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role != Qt.ItemDataRole.DisplayRole:
            return None
        if orientation == Qt.Orientation.Horizontal:
            return self.columns[section]
        return str(section + 1)

    def update(self, columns, rows):
        self.beginResetModel()
        self.columns = columns
        self.rows = rows
        self.endResetModel()

# --------------------- Dynamic dialog for Add/Edit ---------------------
class RecordDialog(QDialog):
    def __init__(self, columns, data=None, parent=None, readonly=None, exclude=None):
        super().__init__(parent)
        self.setWindowTitle('Edit record' if data else 'Add record')
        self.columns = [c for c in columns if not (exclude and c in exclude)]
        self.data = data or {}
        self.fields = {}
        
        layout = QFormLayout()
        layout.setSpacing(12)
        layout.setContentsMargins(20, 20, 20, 20)
        
        font = QFont()
        font.setPointSize(10)
        
        for col in self.columns:
            le = QLineEdit()
            le.setFont(font)
            le.setText(str(self.data.get(col, '')))
            if readonly and col in readonly:
                le.setReadOnly(True)
                le.setStyleSheet("background-color: #f8f9fa;")
            layout.addRow(QLabel(f"{col}:"), le)
            self.fields[col] = le
        
        btn_layout = QHBoxLayout()
        btn_ok = QPushButton('OK')
        btn_ok.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        btn_cancel = QPushButton('Cancel')
        btn_cancel.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        btn_ok.clicked.connect(self.accept)
        btn_cancel.clicked.connect(self.reject)
        btn_layout.addWidget(btn_ok)
        btn_layout.addWidget(btn_cancel)
        layout.addRow(btn_layout)
        
        self.setLayout(layout)
        self.setModal(True)
        self.setMinimumWidth(450)
        
        # Привязка клавиш Enter и Escape
        btn_ok.setDefault(True)
        btn_cancel.setShortcut(QKeySequence(Qt.Key.Key_Escape))
        btn_ok.setShortcut(QKeySequence(Qt.Key.Key_Return))

    def get_data(self):
        return {col: self.fields[col].text() for col in self.columns}

# --------------------- Main Window ---------------------
class MainWindow(QMainWindow):
    TABLE_KEYS = [
        'patients', 'appointments', 'billing', 'departments', 'doctors', 'rooms',
        'doctor_department', 'doctor_patient', 'patient_room', 'medical_records'
    ]
    LOOKUP_KEYS = {'departments', 'rooms'}

    def __init__(self):
        super().__init__()
        self.setWindowTitle('Hospital Admin Client')
        self.resize(1200, 700)
        self.api = ApiClient('http://127.0.0.1:8000')
        self.api.request_finished.connect(self.on_api_result)
        self.last_result = None
        self.current_table = None

        # Apply modern style
        self.apply_style()

        # Create menu bar
        self.create_menu_bar()

        # Create central widget
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setSpacing(0)
        main_layout.setContentsMargins(0, 0, 0, 0)

        # Create toolbar
        self.create_toolbar()
        main_layout.addWidget(self.toolbar)

        # Create table view container
        table_container = QWidget()
        table_layout = QVBoxLayout(table_container)
        table_layout.setContentsMargins(15, 15, 15, 15)
        
        # Create table view
        self.create_table_view()
        table_layout.addWidget(self.table_view)
        
        main_layout.addWidget(table_container)

        # Create status bar
        self.status = QStatusBar()
        self.status.setFont(QFont("Segoe UI", 9))
        self.setStatusBar(self.status)

        # Setup shortcuts
        self.setup_shortcuts()

        # Load default table
        self.cmb_tables.setCurrentText('appointments')
        self.current_table = 'appointments'
        self.load_table(self.current_table)

    def create_menu_bar(self):
        menubar = self.menuBar()
        
        # File menu
        file_menu = menubar.addMenu("&File")  # Alt+F для доступа
        
        # Exit action
        exit_action = QAction("E&xit", self)  # Alt+X для выхода
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # Help menu
        help_menu = menubar.addMenu("&Help")  # Alt+H для доступа
        about_action = QAction("&About", self)  # Alt+A для About
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)

    def show_about(self):
        QMessageBox.about(self, "About", "Hospital Admin Client\n\nA modern interface for hospital database management.")

    def apply_style(self):
        self.setStyleSheet("""
            QMainWindow {
                background-color: #ffffff;
            }

            /* ==========================================
               МЕНЮ БАР (верхнее меню)
            ========================================== */
            QMenuBar {
                background-color: #f8f9fa;
                border-bottom: 1px solid #dee2e6;
                padding: 2px;
                spacing: 6px;
            }

            QMenuBar::item {
                background-color: transparent;
                padding: 6px 12px;
                border-radius: 4px;
                color: #495057; /* Тёмный текст по умолчанию */
                font-size: 12px;
                margin: 0 2px;
            }

            QMenuBar::item:selected {
                background-color: #007bff;
                color: white;
            }

            QMenuBar::item:hover {
                background-color: #e9ecef;
                color: #212529;
            }

            QMenuBar::item:disabled {
                color: #adb5bd;
            }

            /* ==========================================
               МЕНЮ (выпадающие пункты)
            ========================================== */
            QMenu {
                background-color: white;
                border: 1px solid #ced4da;
                border-radius: 6px;
                padding: 4px;
                margin: 2px;
                color: #495057;
                font-size: 12px;
            }

            QMenu::item {
                padding: 6px 24px 6px 12px;
                border-radius: 4px;
                margin: 2px 4px;
            }

            QMenu::item:selected {
                background-color: #007bff;
                color: white;
            }

            QMenu::item:hover {
                background-color: #e9ecef;
                color: #212529;
            }

            QMenu::item:disabled {
                color: #adb5bd;
            }

            QMenu::separator {
                height: 1px;
                background-color: #dee2e6;
                margin: 4px 2px;
            }

            /* ==========================================
               Остальной интерфейс (остальное без изменений)
            ========================================== */
            QToolBar {
                background-color: #f8f9fa;
                padding: 10px;
                spacing: 8px;
                border-bottom: 1px solid #dee2e6;
            }
            QComboBox {
                padding: 6px 12px;
                border: 1px solid #ced4da;
                border-radius: 4px;
                background-color: white;
                min-width: 150px;
                font-size: 12px;
                color: #495057;
            }
            QComboBox:hover {
                border-color: #80bdff;
            }
            QComboBox:focus {
                border-color: #80bdff;
                outline: none;
            }

            /* Выпадающий список QComboBox */
            QComboBox QAbstractItemView {
                border: 1px solid #ced4da;
                selection-background-color: #007bff;
                selection-color: white;
                background-color: white;
                color: #333333;
                font-size: 12px;
                outline: none;
                border-radius: 4px;
            }
            QComboBox QAbstractItemView::item {
                padding: 6px;
                height: 24px;
            }
            QComboBox QAbstractItemView::item:selected {
                background-color: #007bff;
                color: white;
            }
            QComboBox QAbstractItemView::item:hover {
                background-color: #e9ecef;
                color: #333333;
            }

            QPushButton {
                padding: 8px 16px;
                border: 1px solid transparent;
                border-radius: 4px;
                background-color: #007bff;
                color: white;
                font-weight: 500;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #0069d9;
                border-color: #0062cc;
            }
            QPushButton:pressed {
                background-color: #0062cc;
                border-color: #005cbf;
            }
            QPushButton:disabled {
                background-color: #6c757d;
                color: #ffffff;
            }

            QTableView {
                background-color: white;
                alternate-background-color: #f8f9fa;
                gridline-color: #dee2e6;
                border: 1px solid #dee2e6;
                border-radius: 4px;
                font-size: 11px;
            }
            QTableView::item {
                padding: 4px;
                border: none;
                color: #333333;
            }
            QTableView::item:selected {
                background-color: #007bff;
                color: white;
            }
            QTableView::item:hover {
                background-color: #e9ecef;
            }
            QHeaderView::section {
                background-color: #e9ecef;
                padding: 8px;
                border: none;
                font-weight: bold;
                font-size: 11px;
                color: #495057;
            }
            QStatusBar {
                background-color: #f8f9fa;
                color: #6c757d;
                border-top: 1px solid #dee2e6;
                font-size: 11px;
            }
            QLineEdit {
                padding: 6px 12px;
                border: 1px solid #ced4da;
                border-radius: 4px;
                background-color: white;
                font-size: 12px;
                color: #495057;
            }
            QLineEdit:focus {
                border-color: #80bdff;
                outline: none;
            }
            QLabel {
                color: #495057;
                font-size: 12px;
            }
            QDialog {
                background-color: white;
            }
        """)

    def create_toolbar(self):
        self.toolbar = QToolBar()
        self.toolbar.setMovable(False)
        self.toolbar.setIconSize(QSize(16, 16))

        # Table selection
        table_label = QLabel("Table:")
        table_label.setStyleSheet("font-weight: bold; color: #495057;")
        self.toolbar.addWidget(table_label)
        
        self.cmb_tables = QComboBox()
        self.cmb_tables.addItems(self.TABLE_KEYS)
        self.cmb_tables.currentTextChanged.connect(self.on_table_change)
        self.cmb_tables.setMinimumWidth(180)
        self.cmb_tables.setStyleSheet("""
            QComboBox {
                padding: 6px 12px;
                font-size: 12px;
            }
        """)
        self.toolbar.addWidget(self.cmb_tables)

        self.toolbar.addSeparator()

        # Action buttons with improved styling
        buttons_config = [
            ('Reload', QStyle.SP_BrowserReload, "Reload current table (Ctrl+R)", 
             lambda: self.load_table(self.current_table), "#6c757d"),
            ('Add', QStyle.SP_FileDialogNewFolder, "Add new record (Ctrl+A)", 
             self.on_add, "#28a745"),
            ('Edit', QStyle.SP_FileDialogDetailedView, "Edit selected record (Ctrl+U)", 
             self.on_update, "#17a2b8"),
            ('Delete', QStyle.SP_TrashIcon, "Delete selected record (Ctrl+D)", 
             self.on_delete, "#dc3545"),
            ('Backup', QStyle.SP_ComputerIcon, "Create backup (Ctrl+B)", 
             self.on_backup, "#ffc107"),
            ('Restore', QStyle.SP_DriveFDIcon, "Restore from backup", 
             self.on_restore, "#fd7e14"),
            ('Export', QStyle.SP_DialogSaveButton, "Export to CSV", 
             self.on_export, "#20c997"),
            ('Query', QStyle.SP_FileDialogContentsView, "Special query (Ctrl+Q)", 
             self.on_query, "#6610f2"),
            ('Save Excel', QStyle.SP_DialogSaveButton, "Save to Excel (Ctrl+S)", 
             self.on_save_current_table, "#0dcaf0")
        ]

        for text, icon, tooltip, callback, color in buttons_config:
            btn = self.create_styled_button(text, icon, tooltip, callback, color)
            self.toolbar.addWidget(btn)

    def create_styled_button(self, text, icon, tooltip, callback, color):
        btn = QPushButton(text)
        if icon:
            btn.setIcon(self.style().standardIcon(icon))
        btn.setToolTip(tooltip)
        btn.clicked.connect(callback)
        
        btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border: none;
                padding: 8px 12px;
                border-radius: 4px;
                font-weight: 500;
                font-size: 11px;
                margin: 2px;
            }}
            QPushButton:hover {{
                background-color: {self.darken_color(color)};
            }}
            QPushButton:pressed {{
                background-color: {self.darken_color(color, 40)};
            }}
            QPushButton:disabled {{
                background-color: #6c757d;
                color: #adb5bd;
            }}
        """)
        
        return btn

    def darken_color(self, hex_color, percent=20):
        """Darken a hex color by given percentage"""
        hex_color = hex_color.lstrip('#')
        rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        rgb = tuple(max(0, int(c * (100 - percent) / 100)) for c in rgb)
        return f"#{rgb[0]:02x}{rgb[1]:02x}{rgb[2]:02x}"

    def create_table_view(self):
        self.table_view = QTableView()
        self.model = TableModel()
        self.table_view.setModel(self.model)
        self.table_view.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
        self.table_view.setAlternatingRowColors(True)
        self.table_view.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table_view.setSortingEnabled(True)
        self.table_view.setFont(QFont("Segoe UI", 10))
        
        # Improve selection appearance
        self.table_view.setSelectionMode(QTableView.SelectionMode.SingleSelection)
        self.table_view.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)

    def setup_shortcuts(self):
        shortcuts = [
            ('Ctrl+R', lambda: self.load_table(self.current_table)),
            ('Ctrl+A', self.on_add),
            ('Ctrl+U', self.on_update),  # Изменено на Ctrl+U для обновления
            ('Ctrl+D', self.on_delete),
            ('Ctrl+Q', self.on_query),
            ('Ctrl+S', self.on_save_current_table),
            ('Ctrl+B', self.on_backup),
            ('Ctrl+E', self.close),  # Ctrl+E для выхода
            ('F10', self.activate_menu_bar)  # F10 для активации меню
        ]
        
        for key, callback in shortcuts:
            QShortcut(QKeySequence(key), self, activated=callback)

    def activate_menu_bar(self):
        """Активировать меню при нажатии F10"""
        self.menuBar().setFocus()

    def on_save_current_table(self):
        if not self.current_table or not self.last_result:
            QMessageBox.warning(self, "Export", "No data to export")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = self.current_table

        headers = list(self.last_result[0].keys())
        ws.append(headers)

        for row in self.last_result:
            ws.append([row.get(h) for h in headers])

        for col_idx, col in enumerate(headers, 1):
            max_len = max((len(str(r.get(col) or "")) for r in self.last_result), default=len(col))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

        path, _ = QFileDialog.getSaveFileName(
            self, 
            "Save Excel File", 
            f"{self.current_table}.xlsx", 
            "Excel Files (*.xlsx)"
        )
        
        if path:
            try:
                wb.save(path)
                self.status.showMessage(f"✓ Table saved to {path}")
                QMessageBox.information(self, "Success", f"Table successfully saved to:\n{path}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save file:\n{str(e)}")

    def on_table_change(self, key):
        self.current_table = key
        self.load_table(key)

    def load_table(self, key, params=None):
        if not key:
            return
        self.status.showMessage(f'🔄 Loading {key}...')
        self.api.get_table(key, params=params)

    @Slot(object, object)
    def on_api_result(self, result, error):
        if error:
            msg = str(error).strip()
            QMessageBox.critical(self, 'Error', f"API Error:\n{msg}")
            self.status.showMessage('❌ Error')
            return

        if isinstance(result, list):
            self.last_result = result
            columns = list(result[0].keys()) if result else []
            self.model.update(columns, result)
            self.status.showMessage(f'✓ Loaded {len(result)} rows')
        elif isinstance(result, dict):
            self.last_result = result
            if result.get('status') == 'ok':
                self.status.showMessage('✓ Operation successful')
                if self.current_table:
                    self.load_table(self.current_table)
            else:
                msg = json.dumps(result, ensure_ascii=False)
                self.status.showMessage('⚠️ ' + msg)
                QMessageBox.warning(self, 'Response', msg)
        elif isinstance(result, str):
            self.status.showMessage('📝 ' + result)
            QMessageBox.information(self, 'Response', result)
        elif result is None:
            self.status.showMessage('ℹ️ Empty response')
        else:
            self.status.showMessage('❓ Unknown response')

    def selected_row(self):
        sel = self.table_view.selectionModel().selectedRows()
        if not sel:
            return None, None
        row_idx = sel[0].row()
        row = self.model.rows[row_idx]
        return row_idx, row

    def _find_pk(self, row):
        for k in row.keys():
            if k.endswith('_id'):
                return k
        return None

    def ask_and_verify_admin(self, on_success):
        dialog = QDialog(self)
        dialog.setWindowTitle("Admin Authentication")
        dialog.setModal(True)
        dialog.setMinimumWidth(350)
        dialog.setStyleSheet("""
            QDialog {
                background-color: white;
                border-radius: 8px;
            }
        """)
        
        layout = QVBoxLayout(dialog)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        title = QLabel("Admin Authentication Required")
        title.setStyleSheet("font-size: 14px; font-weight: bold; color: #495057;")
        layout.addWidget(title)
        
        form = QFormLayout()
        form.setSpacing(10)
        
        password_edit = QLineEdit()
        password_edit.setEchoMode(QLineEdit.Password)
        password_edit.setPlaceholderText("Enter admin password")
        form.addRow("Password:", password_edit)
        
        layout.addLayout(form)
        
        buttons = QHBoxLayout()
        btn_ok = QPushButton("Verify")
        btn_ok.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        btn_cancel = QPushButton("Cancel")
        btn_cancel.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        buttons.addWidget(btn_ok)
        buttons.addWidget(btn_cancel)
        
        layout.addLayout(buttons)
        
        # Привязка клавиш Enter и Escape
        btn_ok.setDefault(True)
        btn_cancel.setShortcut(QKeySequence(Qt.Key.Key_Escape))
        btn_ok.setShortcut(QKeySequence(Qt.Key.Key_Return))
        
        def verify():
            pw = password_edit.text().strip()
            if not pw:
                QMessageBox.warning(dialog, "Error", "Please enter a password")
                return
                
            self.status.showMessage('🔐 Verifying admin password...')
            
            def verify_callback(result, error):
                if error:
                    QMessageBox.critical(dialog, 'Error', f"Verification failed:\n{error}")
                    self.status.showMessage('❌ Verification failed')
                    return
                if isinstance(result, dict) and result.get('ok'):
                    self.api.set_superpass(pw)
                    self.status.showMessage('✅ Admin verified')
                    dialog.accept()
                    try:
                        on_success()
                    except Exception as e:
                        print('on_success handler error:', e)
                else:
                    msg = result.get('error') if isinstance(result, dict) else 'Invalid password'
                    QMessageBox.warning(dialog, 'Error', f"Authentication failed:\n{msg}")
                    self.status.showMessage('❌ Invalid admin password')

            self.api.verify_superpass(pw, verify_callback)
        
        btn_ok.clicked.connect(verify)
        btn_cancel.clicked.connect(dialog.reject)
        
        dialog.exec()

    def on_add(self):
        def do_add():
            pk_candidates = [c for c in (self.model.columns or []) if c.endswith('_id')]
            dlg_columns = [c for c in (self.model.columns or []) if c not in pk_candidates]
            dlg = RecordDialog(dlg_columns, data=None, parent=self, readonly=None, exclude=None)
            if dlg.exec() == QDialog.Accepted:
                data = dlg.get_data()
                self.api.post(self.current_table, data)
                self.status.showMessage('➕ Adding record...')

        if self.current_table in self.LOOKUP_KEYS:
            self.ask_and_verify_admin(on_success=do_add)
            return

        do_add()

    def on_update(self):
        idx, row = self.selected_row()
        if row is None:
            QMessageBox.information(self, 'Information', 'Please select a row to edit')
            return

        def do_update():
            pk = self._find_pk(row)
            if pk:
                editable_columns = [c for c in self.model.columns if c != pk]
            else:
                editable_columns = list(self.model.columns)
            dlg = RecordDialog(editable_columns, data=row, parent=self, readonly=None, exclude=None)
            if dlg.exec() == QDialog.Accepted:
                data = dlg.get_data()
                if pk:
                    data[pk] = row[pk]
                self.api.put(self.current_table, data)
                self.status.showMessage('✏️ Updating record...')

        if self.current_table in self.LOOKUP_KEYS:
            self.ask_and_verify_admin(on_success=do_update)
            return

        do_update()

    def on_delete(self):
        idx, row = self.selected_row()
        if row is None:
            QMessageBox.information(self, 'Information', 'Please select a row to delete')
            return

        def do_delete():
            reply = QMessageBox.question(
                self, 
                'Confirm Deletion', 
                'Are you sure you want to delete this record?\nThis action cannot be undone.',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                return
            pk = self._find_pk(row)
            if not pk:
                QMessageBox.critical(self, 'Error', 'Primary key not found')
                return
            payload = {pk: row[pk]}
            self.api.delete(self.current_table, payload)
            self.status.showMessage('🗑️ Deleting record...')

        if self.current_table in self.LOOKUP_KEYS:
            self.ask_and_verify_admin(on_success=do_delete)
            return

        do_delete()

    def on_query(self):
        key, ok = QInputDialog.getText(
            self, 
            'Filter Records', 
            'Enter filter criteria (key=value):\nExample: doctor_id=5'
        )
        if not ok or not key:
            return
        try:
            key, val = key.split('=', 1)
            key = key.strip()
            val = val.strip()
        except Exception:
            QMessageBox.warning(self, 'Error', 'Please use format: key=value')
            return
        self.load_table(self.current_table, params={key: val})

    def on_export(self):
        if self.last_result is None:
            QMessageBox.information(self, 'Information', 'No data to export')
            return
        rows = self.last_result if isinstance(self.last_result, list) else []
        if not rows:
            QMessageBox.information(self, 'Information', 'No rows to export')
            return
        
        path, _ = QFileDialog.getSaveFileName(
            self, 
            'Export to CSV', 
            f'{self.current_table}.csv', 
            'CSV files (*.csv)'
        )
        
        if not path:
            return
            
        cols = list(rows[0].keys())
        try:
            with open(path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=cols)
                writer.writeheader()
                for r in rows:
                    writer.writerow(r)
            self.status.showMessage(f'📊 Exported {len(rows)} rows to {path}')
            QMessageBox.information(self, 'Success', f'Successfully exported {len(rows)} rows to:\n{path}')
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Export failed:\n{str(e)}')

    def on_backup(self):
        reply = QMessageBox.question(
            self, 
            'Create Backup', 
            'Create a backup of the database now?'
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        self.ask_and_verify_admin(on_success=self.api.backup)

    def on_restore(self):
        reply = QMessageBox.question(
            self, 
            'Restore Database', 
            'WARNING: This will overwrite all current data!\nRestore database from backup?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        self.ask_and_verify_admin(on_success=self.api.restore)

# --------------------- Entry point ---------------------
if __name__ == '__main__':
    app = QApplication(sys.argv)
    
    # Set application style and font
    app.setStyle('Fusion')
    
    # Set application font
    font = QFont("Segoe UI", 10)
    app.setFont(font)
    
    # Create and show main window
    win = MainWindow()
    win.show()
    
    sys.exit(app.exec())